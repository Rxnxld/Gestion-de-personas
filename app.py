import os, csv, io, hashlib, secrets
from datetime import timedelta
from flask import Flask, g, request, jsonify, send_from_directory, session, Response, send_file
from flask_cors import CORS
import psycopg2
import psycopg2.extras

app = Flask(__name__, static_folder='.', static_url_path='')
# Usa una SECRET_KEY fija (variable de entorno en Render) para que las sesiones
# no se invaliden cada vez que el servidor reinicia o se redespliega.
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.permanent_session_lifetime = timedelta(hours=8)
CORS(app, supports_credentials=True)

# En Render, crea una base de datos PostgreSQL y copia su "Internal Database URL"
# como variable de entorno DATABASE_URL en este servicio web.
DATABASE_URL = os.environ['DATABASE_URL']

class DBWrapper:
    """Envuelve psycopg2 para que el resto del código (escrito para sqlite3)
    siga funcionando igual: placeholders '?' y filas accesibles por nombre."""
    def __init__(self, conn):
        self.conn = conn
    def execute(self, sql, params=()):
        cur = self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql.replace('?', '%s'), params)
        return cur
    def commit(self):
        self.conn.commit()
    def close(self):
        self.conn.close()

def get_db():
    if 'db' not in g:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        g.db = DBWrapper(conn)
    return g.db

def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()
app.teardown_appcontext(close_db)

# ═══════════════════════════════════════════
#  CONFIG (password) guardada en la BD, no en archivo
# ═══════════════════════════════════════════

def _config_conn():
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = True
    return conn

def get_config_value(key):
    conn = _config_conn()
    cur = conn.cursor()
    cur.execute("SELECT value FROM app_config WHERE key=%s", (key,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None

def set_config_value(key, value):
    conn = _config_conn()
    cur = conn.cursor()
    cur.execute("""INSERT INTO app_config (key, value) VALUES (%s, %s)
                   ON CONFLICT (key) DO UPDATE SET value = excluded.value""", (key, value))
    conn.close()

def init_config():
    if get_config_value('PASSWORD') is None:
        pw = 'admin123'
        set_config_value('PASSWORD', hashlib.sha256(pw.encode()).hexdigest())
        print(f"  ── Config creada: password por defecto = {pw}")

def check_password(pw):
    stored = get_config_value('PASSWORD')
    if not stored: return False
    return stored == hashlib.sha256(pw.encode()).hexdigest()

def set_password(pw):
    set_config_value('PASSWORD', hashlib.sha256(pw.encode()).hexdigest())

def login_required(fn):
    def wrapper(*a, **kw):
        if not session.get('logged_in'):
            return jsonify({'error':'No autenticado'}), 401
        return fn(*a, **kw)
    wrapper.__name__ = fn.__name__
    return wrapper

def init_db():
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = True
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS app_config (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS miembros (
            id SERIAL PRIMARY KEY,
            nombre TEXT UNIQUE NOT NULL,
            apodo TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS fechas_tablas (
            id SERIAL PRIMARY KEY,
            fecha TEXT UNIQUE NOT NULL
        );
        CREATE TABLE IF NOT EXISTS asistencias (
            miembro_id INTEGER NOT NULL REFERENCES miembros(id) ON DELETE CASCADE,
            fecha TEXT NOT NULL,
            valor INTEGER DEFAULT 0,
            PRIMARY KEY (miembro_id, fecha)
        );
        CREATE TABLE IF NOT EXISTS bingos (
            id SERIAL PRIMARY KEY,
            fecha TEXT UNIQUE NOT NULL,
            monto REAL NOT NULL,
            adicional REAL DEFAULT 0,
            asistentes INTEGER DEFAULT 0
        );
        ALTER TABLE bingos ADD COLUMN IF NOT EXISTS asistentes INTEGER DEFAULT 0;
        ALTER TABLE bingos ADD COLUMN IF NOT EXISTS adicional REAL DEFAULT 0;
        CREATE TABLE IF NOT EXISTS bingo_distribucion (
            id SERIAL PRIMARY KEY,
            bingo_id INTEGER NOT NULL REFERENCES bingos(id) ON DELETE CASCADE,
            miembro_id INTEGER NOT NULL REFERENCES miembros(id) ON DELETE CASCADE,
            recibe BOOLEAN DEFAULT TRUE,
            monto_asignado REAL DEFAULT 0,
            personalizado BOOLEAN DEFAULT FALSE,
            UNIQUE(bingo_id, miembro_id)
        );
        -- backfill: crea en bingos las fechas que ya existen en fechas_tablas
        -- pero que todavia no tienen fila espejo (p.ej. fechas creadas antes
        -- de conectar ambas tablas)
        INSERT INTO bingos (fecha, monto, adicional, asistentes)
        SELECT ft.fecha, 0, 0, 0
        FROM fechas_tablas ft
        WHERE NOT EXISTS (SELECT 1 FROM bingos b WHERE b.fecha = ft.fecha)
        ON CONFLICT (fecha) DO NOTHING;
        -- backfill: crea en fechas_tablas las fechas que ya existen en bingos
        -- pero que todavia no tienen fila espejo (por si fue al reves)
        INSERT INTO fechas_tablas (fecha)
        SELECT b.fecha
        FROM bingos b
        WHERE NOT EXISTS (SELECT 1 FROM fechas_tablas ft WHERE ft.fecha = b.fecha)
        ON CONFLICT (fecha) DO NOTHING;
        -- backfill distribucion para bingos existentes sin distribucion
        -- (incluye los que se acaban de crear en el paso anterior).
        -- Si esa fecha ya tiene asistencia registrada en Tablas de Asistencia,
        -- respeta quien asistio de verdad; si no hay asistencia registrada
        -- para esa fecha (bingos antiguos sin tabla asociada), usa TRUE por
        -- defecto para no romper datos historicos.
        INSERT INTO bingo_distribucion (bingo_id, miembro_id, recibe, monto_asignado)
        SELECT b.id, m.id,
               CASE WHEN EXISTS (SELECT 1 FROM asistencias a2 WHERE a2.fecha = b.fecha)
                    THEN COALESCE(a.valor, 0) = 1
                    ELSE TRUE END AS recibe,
               CASE WHEN EXISTS (SELECT 1 FROM asistencias a2 WHERE a2.fecha = b.fecha)
                        AND COALESCE(a.valor, 0) != 1
                    THEN 0
                    ELSE ROUND(((b.monto + COALESCE(b.adicional,0)) / NULLIF(b.asistentes,0))::numeric, 2)::real
               END AS monto_asignado
        FROM bingos b
        CROSS JOIN miembros m
        LEFT JOIN asistencias a ON a.miembro_id = m.id AND a.fecha = b.fecha
        WHERE NOT EXISTS (SELECT 1 FROM bingo_distribucion d WHERE d.bingo_id=b.id AND d.miembro_id=m.id)
        ON CONFLICT (bingo_id, miembro_id) DO NOTHING;
        -- re-sincroniza recibe/monto en cada arranque para bingos cuya fecha
        -- ya tenga asistencia registrada (por si se marco/edito asistencia
        -- despues de crear el bingo, lo cual antes dejaba el reparto
        -- desactualizado). No toca reparto personalizado.
        UPDATE bingo_distribucion d
        SET recibe = (
              SELECT COALESCE(a.valor, 0) = 1
              FROM asistencias a
              WHERE a.miembro_id = d.miembro_id AND a.fecha = b.fecha
            ),
            monto_asignado = CASE WHEN (
                SELECT COALESCE(a.valor, 0) = 1
                FROM asistencias a
                WHERE a.miembro_id = d.miembro_id AND a.fecha = b.fecha
              )
              THEN COALESCE(ROUND(((b.monto + COALESCE(b.adicional,0)) / NULLIF(b.asistentes,0))::numeric, 2)::real, 0)
              ELSE 0
            END
        FROM bingos b
        WHERE d.bingo_id = b.id
          AND d.personalizado = FALSE
          AND EXISTS (SELECT 1 FROM asistencias a2 WHERE a2.fecha = b.fecha);
        CREATE TABLE IF NOT EXISTS fechas_rifa (
            id SERIAL PRIMARY KEY,
            fecha TEXT UNIQUE NOT NULL
        );
        CREATE TABLE IF NOT EXISTS rifas (
            miembro_id INTEGER NOT NULL REFERENCES miembros(id) ON DELETE CASCADE,
            fecha TEXT NOT NULL,
            valor INTEGER DEFAULT 0,
            PRIMARY KEY (miembro_id, fecha)
        );
        CREATE TABLE IF NOT EXISTS meses_ahorro (
            id INTEGER PRIMARY KEY,
            nombre TEXT UNIQUE NOT NULL,
            orden INTEGER NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ahorros (
            id SERIAL PRIMARY KEY,
            tipo TEXT NOT NULL CHECK(tipo IN ('normal','cumple','rifa')),
            miembro_id INTEGER NOT NULL REFERENCES miembros(id) ON DELETE CASCADE,
            mes_id INTEGER NOT NULL REFERENCES meses_ahorro(id) ON DELETE CASCADE,
            valor REAL DEFAULT 0,
            UNIQUE(tipo, miembro_id, mes_id)
        );
        CREATE TABLE IF NOT EXISTS cumple_meses (
            id INTEGER PRIMARY KEY,
            clave TEXT UNIQUE NOT NULL,
            orden INTEGER NOT NULL
        );
        CREATE TABLE IF NOT EXISTS cumple_aportes (
            cumple_mes_id INTEGER NOT NULL REFERENCES cumple_meses(id) ON DELETE CASCADE,
            miembro_id INTEGER NOT NULL REFERENCES miembros(id) ON DELETE CASCADE,
            valor INTEGER DEFAULT 0,
            PRIMARY KEY (cumple_mes_id, miembro_id)
        );
        CREATE TABLE IF NOT EXISTS cumple_fechas (
            miembro_id INTEGER PRIMARY KEY REFERENCES miembros(id) ON DELETE CASCADE,
            fecha TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS prestamos (
            id SERIAL PRIMARY KEY,
            miembro_id INTEGER NOT NULL REFERENCES miembros(id) ON DELETE CASCADE,
            monto REAL NOT NULL,
            obs TEXT DEFAULT '',
            estado TEXT NOT NULL DEFAULT 'pendiente' CHECK (estado IN ('pendiente','pagado')),
            fecha TEXT DEFAULT NULL
        );
        ALTER TABLE prestamos ADD COLUMN IF NOT EXISTS pagado REAL DEFAULT 0;
        CREATE TABLE IF NOT EXISTS abonos (
            id SERIAL PRIMARY KEY,
            prestamo_id INTEGER NOT NULL REFERENCES prestamos(id) ON DELETE CASCADE,
            monto REAL NOT NULL,
            fecha TEXT DEFAULT NULL
        );
        INSERT INTO meses_ahorro (id, nombre, orden) VALUES
            (1,'ENERO',1),(2,'FEBRERO',2),(3,'MARZO',3),(4,'ABRIL',4),
            (5,'MAYO',5),(6,'JUNIO',6),(7,'JULIO',7),(8,'AGOSTO',8),
            (9,'SEPTIEMBRE',9),(10,'OCTUBRE',10),(11,'NOVIEMBRE',11),(12,'DICIEMBRE',12)
        ON CONFLICT (id) DO NOTHING;
        INSERT INTO cumple_meses (id, clave, orden) VALUES
            (1,'ENERO (2)',1),(2,'FEBRERO (4)',2),(3,'MARZO(2)',3),
            (4,'ABRIL(1)',4),(5,'MAYO(2)',5),(6,'OCTUBRE(4)',6),(7,'DICIEMBRE(3)',7)
        ON CONFLICT (id) DO NOTHING;
    ''')
    # Migraciones para bases ya existentes (columnas nuevas en prestamos)
    c.execute("ALTER TABLE prestamos ADD COLUMN IF NOT EXISTS estado TEXT NOT NULL DEFAULT 'pendiente'")
    c.execute("ALTER TABLE prestamos ADD COLUMN IF NOT EXISTS fecha TEXT DEFAULT NULL")
    # Marca si el monto de un miembro en un bingo fue personalizado a mano en el
    # modal de Reparto; si es así, los recálculos automáticos no deben pisarlo.
    c.execute("ALTER TABLE bingo_distribucion ADD COLUMN IF NOT EXISTS personalizado BOOLEAN DEFAULT FALSE")
    conn.close()

# ═══════════════════════════════════════════
#  AUTH
# ═══════════════════════════════════════════

@app.route('/api/auth/status')
def auth_status():
    configurado = get_config_value('PASSWORD') is not None
    return jsonify({'logged_in': session.get('logged_in', False), 'configurado': configurado})

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    pw = request.json.get('password', '')
    if not check_password(pw):
        return jsonify({'error':'Contraseña incorrecta'}), 401
    session.permanent = True
    session['logged_in'] = True
    return jsonify({'ok': True})

@app.route('/api/auth/setup', methods=['POST'])
def auth_setup():
    if get_config_value('PASSWORD') is not None and not session.get('logged_in'):
        return jsonify({'error':'Ya configurado. Inicia sesión primero.'}), 400
    pw = request.json.get('password', '')
    if len(pw) < 4: return jsonify({'error':'Mínimo 4 caracteres'}), 400
    set_password(pw)
    session.permanent = True
    session['logged_in'] = True
    return jsonify({'ok': True})

@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def auth_change():
    old = request.json.get('old', '')
    new = request.json.get('new', '')
    if not check_password(old): return jsonify({'error':'Contraseña actual incorrecta'}), 400
    if len(new) < 4: return jsonify({'error':'Mínimo 4 caracteres'}), 400
    set_password(new)
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
#  MIEMBROS
# ═══════════════════════════════════════════

@app.route('/api/miembros', methods=['GET'])
@login_required
def get_miembros():
    return jsonify([dict(r) for r in get_db().execute("SELECT id, nombre, apodo FROM miembros ORDER BY id").fetchall()])

@app.route('/api/miembros', methods=['POST'])
@login_required
def add_miembro():
    data = request.json
    nombre = data.get('nombre', '').strip().upper()
    if not nombre: return jsonify({'error':'Nombre requerido'}), 400
    try:
        c = get_db().execute("INSERT INTO miembros (nombre, apodo) VALUES (?, ?) RETURNING id", (nombre, data.get('apodo','').strip()))
        new_id = c.fetchone()['id']
        get_db().commit()
        return jsonify({'id': new_id, 'nombre': nombre}), 201
    except psycopg2.IntegrityError:
        return jsonify({'error':'El miembro ya existe'}), 409

@app.route('/api/miembros/<int:id>', methods=['DELETE'])
@login_required
def delete_miembro(id):
    get_db().execute("DELETE FROM miembros WHERE id=?", (id,))
    get_db().commit()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
#  TABLAS / ASISTENCIAS
# ═══════════════════════════════════════════

@app.route('/api/tablas/fechas', methods=['GET'])
@login_required
def get_fechas_tablas():
    return jsonify([r['fecha'] for r in get_db().execute("SELECT fecha FROM fechas_tablas ORDER BY fecha").fetchall()])

@app.route('/api/tablas/fechas', methods=['POST'])
@login_required
def add_fecha_tablas():
    fecha = request.json.get('fecha')
    if not fecha: return jsonify({'error':'Fecha requerida'}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO fechas_tablas (fecha) VALUES (?)", (fecha,))
    except psycopg2.IntegrityError:
        return jsonify({'error':'La fecha ya existe'}), 409

    # Conecta con Bingo: crea también la fecha en la tabla de bingos
    # (si aún no existe) para que ambas tablas queden sincronizadas.
    cur = db.execute(
        "INSERT INTO bingos (fecha, monto, adicional, asistentes) VALUES (?, 0, 0, 0) "
        "ON CONFLICT (fecha) DO NOTHING RETURNING id", (fecha,))
    row = cur.fetchone()
    if row:
        bingo_id = row['id']
        attendees = {r['miembro_id'] for r in db.execute(
            "SELECT miembro_id FROM asistencias WHERE fecha=? AND valor=1", (fecha,)).fetchall()}
        for m in db.execute("SELECT id FROM miembros").fetchall():
            db.execute(
                "INSERT INTO bingo_distribucion (bingo_id, miembro_id, recibe, monto_asignado) "
                "VALUES (?, ?, ?, 0) ON CONFLICT (bingo_id, miembro_id) DO NOTHING",
                (bingo_id, m['id'], m['id'] in attendees))

    db.commit()
    return jsonify({'ok': True}), 201

@app.route('/api/tablas/asistencias', methods=['GET'])
@login_required
def get_asistencias():
    rows = get_db().execute("SELECT m.nombre, a.fecha, a.valor FROM asistencias a JOIN miembros m ON a.miembro_id = m.id").fetchall()
    r = {}
    for row in rows:
        if row['nombre'] not in r: r[row['nombre']] = {}
        r[row['nombre']][row['fecha']] = row['valor']
    return jsonify(r)

@app.route('/api/tablas/asistencias', methods=['POST'])
@login_required
def set_asistencia():
    data = request.json
    db = get_db()
    miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (data['nombre'],)).fetchone()
    if not miembro: return jsonify({'error':'Miembro no encontrado'}), 404
    valor = int(data.get('valor', 0))
    db.execute("INSERT INTO asistencias (miembro_id, fecha, valor) VALUES (?, ?, ?) ON CONFLICT(miembro_id, fecha) DO UPDATE SET valor=excluded.valor",
               (miembro['id'], data['fecha'], valor))
    # Conecta con Bingo: si esa fecha ya tiene un bingo asociado, sincroniza
    # si el miembro "recibe" reparto segun la asistencia recien marcada y
    # recalcula su monto (a menos que ya tenga un reparto personalizado).
    bingo = db.execute("SELECT id, monto, adicional, asistentes FROM bingos WHERE fecha=?", (data['fecha'],)).fetchone()
    if bingo:
        recibe = valor == 1
        total = (bingo['monto'] or 0) + (bingo['adicional'] or 0)
        a = bingo['asistentes'] or 0
        coge = round(total / a, 2) if a > 0 else 0
        db.execute(
            "UPDATE bingo_distribucion SET recibe=?, monto_asignado=? "
            "WHERE bingo_id=? AND miembro_id=? AND personalizado=FALSE",
            (recibe, coge if recibe else 0, bingo['id'], miembro['id']))
    db.commit()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
#  BINGOS
# ═══════════════════════════════════════════

@app.route('/api/bingos', methods=['GET'])
@login_required
def get_bingos():
    db = get_db()
    bingos = [dict(r) for r in db.execute("SELECT id, fecha, monto, adicional, asistentes FROM bingos ORDER BY fecha").fetchall()]
    for b in bingos:
        dist = db.execute("SELECT miembro_id, recibe, monto_asignado, personalizado FROM bingo_distribucion WHERE bingo_id=?", (b["id"],)).fetchall()
        b['distribucion'] = [dict(d) for d in dist] if dist else []
    return jsonify(bingos)

@app.route('/api/bingos', methods=['POST'])
@login_required
def add_bingo():
    data = request.json
    db = get_db()
    adicional = float(data.get('adicional', 0) or 0)
    try:
        cur = db.execute("INSERT INTO bingos (fecha, monto, adicional, asistentes) VALUES (?, ?, ?, ?) RETURNING id",
                         (data['fecha'], float(data['monto']), adicional, int(data.get('asistentes', 0))))
        bingo_id = cur.fetchone()['id']
        # Quiénes asistieron ese día (asistencias con valor=1)
        attendees = {r['miembro_id'] for r in db.execute(
            "SELECT miembro_id FROM asistencias WHERE fecha=? AND valor=1", (data['fecha'],)).fetchall()}
        asistentes = len(attendees) or int(data.get('asistentes', 0))
        total = float(data['monto']) + adicional
        coge = round(total / asistentes, 2) if asistentes > 0 else 0
        todos = db.execute("SELECT id FROM miembros").fetchall()
        for m in todos:
            recibe = m['id'] in attendees
            db.execute("INSERT INTO bingo_distribucion (bingo_id, miembro_id, recibe, monto_asignado) VALUES (?, ?, ?, ?)",
                       (bingo_id, m['id'], recibe, coge if recibe else 0))
        try:
            db.execute("INSERT INTO fechas_tablas (fecha) VALUES (?)", (data['fecha'],))
        except psycopg2.IntegrityError:
            pass
        db.commit()
        return jsonify({'ok': True}), 201
    except psycopg2.IntegrityError:
        return jsonify({'error': 'La fecha ya existe'}), 409

@app.route('/api/bingos/<int:id>', methods=['PUT'])
@login_required
def update_bingo(id):
    data = request.json
    db = get_db()
    old = db.execute("SELECT id, fecha, monto, adicional, asistentes FROM bingos WHERE id=?", (id,)).fetchone()
    if 'asistentes' in data:
        db.execute("UPDATE bingos SET asistentes=? WHERE id=?", (int(data['asistentes']), id))
    if 'monto' in data:
        db.execute("UPDATE bingos SET monto=? WHERE id=?", (float(data['monto']), id))
    if 'adicional' in data:
        db.execute("UPDATE bingos SET adicional=? WHERE id=?", (float(data['adicional']), id))
    if 'fecha' in data and old and data['fecha'] != old['fecha']:
        try:
            db.execute("UPDATE bingos SET fecha=? WHERE id=?", (data['fecha'], id))
            db.execute("UPDATE fechas_tablas SET fecha=? WHERE fecha=?", (data['fecha'], old['fecha']))
            db.execute("UPDATE asistencias SET fecha=? WHERE fecha=?", (data['fecha'], old['fecha']))
        except psycopg2.IntegrityError:
            return jsonify({'error': 'La fecha ya existe en bingos'}), 409
    # si cambió monto/adicional/asistentes, actualizar distribución automática
    if 'monto' in data or 'adicional' in data or 'asistentes' in data:
        b = db.execute("SELECT monto, adicional, asistentes FROM bingos WHERE id=?", (id,)).fetchone()
        if b:
            total = b['monto'] + (b['adicional'] or 0)
            a = b['asistentes'] or 1
            coge = round(total / a, 2) if a > 0 else 0
            db.execute("UPDATE bingo_distribucion SET monto_asignado=? WHERE bingo_id=? AND recibe=TRUE AND personalizado=FALSE", (coge, id))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/bingos/<int:id>', methods=['DELETE'])
@login_required
def delete_bingo(id):
    db = get_db()
    row = db.execute("SELECT fecha FROM bingos WHERE id=?", (id,)).fetchone()
    db.execute("DELETE FROM bingos WHERE id=?", (id,))
    # Mantiene sincronizada Tablas de Asistencia: si esa fecha no tiene
    # asistencias registradas, se elimina también de fechas_tablas.
    if row:
        tiene_asistencias = db.execute(
            "SELECT 1 FROM asistencias WHERE fecha=? LIMIT 1", (row['fecha'],)).fetchone()
        if not tiene_asistencias:
            db.execute("DELETE FROM fechas_tablas WHERE fecha=?", (row['fecha'],))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/bingos/<int:id>/distribucion', methods=['GET'])
@login_required
def get_distribucion(id):
    db = get_db()
    dist = {r['miembro_id']: {'recibe': r['recibe'], 'monto': r['monto_asignado'], 'personalizado': r['personalizado']}
            for r in db.execute("SELECT miembro_id, recibe, monto_asignado, personalizado FROM bingo_distribucion WHERE bingo_id=?", (id,)).fetchall()}
    miembros = db.execute("SELECT id, nombre FROM miembros ORDER BY nombre").fetchall()
    return jsonify([{'miembro_id': m['id'], 'nombre': m['nombre'],
                     'recibe': dist.get(m['id'], {}).get('recibe', False),
                     'monto': dist.get(m['id'], {}).get('monto', 0),
                     'personalizado': dist.get(m['id'], {}).get('personalizado', False)} for m in miembros])

@app.route('/api/bingos/<int:id>/distribucion', methods=['PUT'])
@login_required
def save_distribucion(id):
    data = request.json
    db = get_db()
    b = db.execute("SELECT monto, adicional, asistentes FROM bingos WHERE id=?", (id,)).fetchone()
    a = (b['asistentes'] or 1) if b else 1
    coge = round(((b['monto'] + (b['adicional'] or 0)) / a), 2) if b else 0
    db.execute("DELETE FROM bingo_distribucion WHERE bingo_id=?", (id,))
    for item in data:
        recibe = item.get('recibe', True)
        if not recibe:
            monto, personalizado = 0, False
        else:
            enviado = round(float(item.get('monto', coge)), 2)
            # Si el monto que llega es distinto al "Coge c/u" automático, es que
            # lo personalizaron a mano (ganó un extra ese día, etc.). Se marca
            # como personalizado para que los recálculos automáticos (cambios
            # de asistencia, monto, adicional o asistentes) no lo pisen.
            personalizado = abs(enviado - coge) > 0.001
            monto = enviado if personalizado else coge
        db.execute("INSERT INTO bingo_distribucion (bingo_id, miembro_id, recibe, monto_asignado, personalizado) VALUES (?, ?, ?, ?, ?)",
                   (id, item['miembro_id'], recibe, monto, personalizado))
    db.commit()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
#  RIFA
# ═══════════════════════════════════════════

@app.route('/api/rifa/fechas', methods=['GET'])
@login_required
def get_fechas_rifa():
    return jsonify([r['fecha'] for r in get_db().execute("SELECT fecha FROM fechas_rifa ORDER BY fecha").fetchall()])

@app.route('/api/rifa/fechas', methods=['POST'])
@login_required
def add_fecha_rifa():
    fecha = request.json.get('fecha')
    if not fecha: return jsonify({'error':'Fecha requerida'}), 400
    try:
        get_db().execute("INSERT INTO fechas_rifa (fecha) VALUES (?)", (fecha,))
        get_db().commit()
        return jsonify({'ok': True}), 201
    except psycopg2.IntegrityError:
        return jsonify({'error':'La fecha ya existe'}), 409

@app.route('/api/rifa', methods=['GET'])
@login_required
def get_rifas():
    rows = get_db().execute("SELECT m.nombre, r.fecha as mes, r.valor FROM rifas r JOIN miembros m ON r.miembro_id = m.id").fetchall()
    r = {}
    for row in rows:
        if row['nombre'] not in r: r[row['nombre']] = {}
        r[row['nombre']][row['mes']] = row['valor']
    return jsonify(r)

@app.route('/api/rifa', methods=['POST'])
@login_required
def set_rifa():
    data = request.json
    db = get_db()
    miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (data['nombre'],)).fetchone()
    if not miembro: return jsonify({'error':'Miembro no encontrado'}), 404
    valor = int(data.get('valor', 0))
    if valor < 0 or valor > 3: return jsonify({'error':'Valor debe ser 0, 1, 2 o 3'}), 400
    db.execute("INSERT INTO rifas (miembro_id, fecha, valor) VALUES (?, ?, ?) ON CONFLICT(miembro_id, fecha) DO UPDATE SET valor=excluded.valor",
               (miembro['id'], data['mes'], valor))
    db.commit()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
#  AHORROS
# ═══════════════════════════════════════════

@app.route('/api/ahorros/<tipo>', methods=['GET'])
@login_required
def get_ahorros(tipo):
    if tipo not in ('normal','cumple','rifa'): return jsonify({'error':'Tipo inválido'}), 400
    rows = get_db().execute("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo=?", (tipo,)).fetchall()
    r = {}
    for row in rows:
        if row['nombre'] not in r: r[row['nombre']] = {}
        r[row['nombre']][row['mes']] = row['valor']
    return jsonify(r)

@app.route('/api/ahorros/<tipo>', methods=['POST'])
@login_required
def set_ahorro(tipo):
    if tipo not in ('normal','cumple','rifa'): return jsonify({'error':'Tipo inválido'}), 400
    data = request.json
    db = get_db()
    miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (data['nombre'],)).fetchone()
    if not miembro: return jsonify({'error':'Miembro no encontrado'}), 404
    mes_row = db.execute("SELECT id FROM meses_ahorro WHERE nombre=?", (data['mes'],)).fetchone()
    if not mes_row: return jsonify({'error':'Mes inválido'}), 400
    db.execute("INSERT INTO ahorros (tipo, miembro_id, mes_id, valor) VALUES (?, ?, ?, ?) ON CONFLICT(tipo, miembro_id, mes_id) DO UPDATE SET valor=excluded.valor",
               (tipo, miembro['id'], mes_row['id'], float(data.get('valor', 0))))
    db.commit()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
#  CUMPLEAÑOS
# ═══════════════════════════════════════════

@app.route('/api/cumple/meses', methods=['GET'])
@login_required
def get_cumple_meses():
    return jsonify([r['clave'] for r in get_db().execute("SELECT clave FROM cumple_meses ORDER BY orden").fetchall()])

@app.route('/api/cumple/aportes', methods=['GET'])
@login_required
def get_cumple_aportes():
    rows = get_db().execute("SELECT cm.clave, m.nombre, ca.valor FROM cumple_aportes ca JOIN cumple_meses cm ON ca.cumple_mes_id = cm.id JOIN miembros m ON ca.miembro_id = m.id").fetchall()
    r = {}
    for row in rows:
        if row['clave'] not in r: r[row['clave']] = {}
        r[row['clave']][row['nombre']] = row['valor']
    return jsonify(r)

@app.route('/api/cumple/aportes', methods=['POST'])
@login_required
def set_cumple_aporte():
    data = request.json
    db = get_db()
    miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (data['nombre'],)).fetchone()
    if not miembro: return jsonify({'error':'Miembro no encontrado'}), 404
    cm = db.execute("SELECT id FROM cumple_meses WHERE clave=?", (data['clave'],)).fetchone()
    if not cm: return jsonify({'error':'Mes inválido'}), 400
    db.execute("INSERT INTO cumple_aportes (cumple_mes_id, miembro_id, valor) VALUES (?, ?, ?) ON CONFLICT(cumple_mes_id, miembro_id) DO UPDATE SET valor=excluded.valor",
               (cm['id'], miembro['id'], int(data.get('valor', 0))))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/cumple/fechas', methods=['GET'])
@login_required
def get_cumple_fechas():
    rows = get_db().execute("SELECT m.nombre, cf.fecha FROM cumple_fechas cf JOIN miembros m ON cf.miembro_id = m.id").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/cumple/fechas', methods=['POST'])
@login_required
def add_cumple_fecha():
    data = request.json
    miembro = get_db().execute("SELECT id FROM miembros WHERE nombre=?", (data['nombre'],)).fetchone()
    if not miembro: return jsonify({'error':'Miembro no encontrado'}), 404
    try:
        get_db().execute("INSERT INTO cumple_fechas (miembro_id, fecha) VALUES (?, ?)", (miembro['id'], data['fecha']))
        get_db().commit()
        return jsonify({'ok': True}), 201
    except psycopg2.IntegrityError:
        return jsonify({'error':'El miembro ya tiene fecha'}), 409

@app.route('/api/cumple/fechas/<int:id>', methods=['DELETE'])
@login_required
def delete_cumple_fecha(id):
    get_db().execute("DELETE FROM cumple_fechas WHERE miembro_id=?", (id,))
    get_db().commit()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
#  PRESTAMOS
# ═══════════════════════════════════════════

@app.route('/api/prestamos', methods=['GET'])
@login_required
def get_prestamos():
    return jsonify([dict(r) for r in get_db().execute(
        "SELECT p.id, m.nombre, p.monto, p.obs, p.estado, p.fecha, COALESCE(p.pagado,0) as pagado "
        "FROM prestamos p JOIN miembros m ON p.miembro_id = m.id ORDER BY p.id").fetchall()])

@app.route('/api/prestamos', methods=['POST'])
@login_required
def add_prestamo():
    data = request.json
    miembro = get_db().execute("SELECT id FROM miembros WHERE nombre=?", (data['nombre'],)).fetchone()
    if not miembro: return jsonify({'error':'Miembro no encontrado'}), 404
    c = get_db().execute("INSERT INTO prestamos (miembro_id, monto, obs, fecha) VALUES (?, ?, ?, ?) RETURNING id",
                          (miembro['id'], float(data['monto']), data.get('obs', ''), data.get('fecha') or None))
    new_id = c.fetchone()['id']
    get_db().commit()
    return jsonify({'id': new_id}), 201

@app.route('/api/prestamos/<int:id>', methods=['DELETE'])
@login_required
def delete_prestamo(id):
    get_db().execute("DELETE FROM prestamos WHERE id=?", (id,))
    get_db().commit()
    return jsonify({'ok': True})

@app.route('/api/prestamos/<int:id>/estado', methods=['POST'])
@login_required
def toggle_prestamo_estado(id):
    db = get_db()
    row = db.execute("SELECT estado, monto FROM prestamos WHERE id=?", (id,)).fetchone()
    if not row: return jsonify({'error':'Préstamo no encontrado'}), 404
    nuevo = 'pagado' if row['estado'] == 'pendiente' else 'pendiente'
    if nuevo == 'pagado':
        db.execute("UPDATE prestamos SET estado=?, pagado=monto WHERE id=?", (nuevo, id))
    else:
        db.execute("UPDATE prestamos SET estado=?, pagado=0 WHERE id=?", (nuevo, id))
    db.commit()
    return jsonify({'ok': True, 'estado': nuevo})

@app.route('/api/prestamos/<int:id>/abono', methods=['POST'])
@login_required
def add_abono(id):
    data = request.json
    db = get_db()
    row = db.execute("SELECT id, monto, COALESCE(pagado,0) as pagado FROM prestamos WHERE id=?", (id,)).fetchone()
    if not row: return jsonify({'error':'Préstamo no encontrado'}), 404
    monto = float(data['monto'])
    if monto <= 0: return jsonify({'error':'Monto debe ser positivo'}), 400
    nuevo_pagado = row['pagado'] + monto
    if nuevo_pagado > row['monto']: return jsonify({'error':'El abono excede la deuda pendiente'}), 400
    db.execute("INSERT INTO abonos (prestamo_id, monto, fecha) VALUES (?, ?, ?)",
               (row['id'], monto, data.get('fecha') or None))
    nuevo_estado = 'pagado' if nuevo_pagado >= row['monto'] else 'pendiente'
    db.execute("UPDATE prestamos SET pagado=?, estado=? WHERE id=?", (nuevo_pagado, nuevo_estado, row['id']))
    db.commit()
    return jsonify({'ok': True, 'saldo': round(row['monto'] - nuevo_pagado, 2), 'estado': nuevo_estado})

@app.route('/api/prestamos/<int:id>/abonos', methods=['GET'])
@login_required
def get_abonos(id):
    abonos = [dict(r) for r in get_db().execute("SELECT id, monto, fecha FROM abonos WHERE prestamo_id=? ORDER BY id", (id,)).fetchall()]
    return jsonify(abonos)

@app.route('/api/abonos/<int:id>', methods=['PUT'])
@login_required
def update_abono(id):
    data = request.json
    db = get_db()
    row = db.execute("SELECT a.id, a.prestamo_id, a.monto FROM abonos a WHERE a.id=?", (id,)).fetchone()
    if not row: return jsonify({'error':'Abono no encontrado'}), 404
    nuevo_monto = float(data['monto'])
    if nuevo_monto <= 0: return jsonify({'error':'Monto debe ser positivo'}), 400
    prestamo = db.execute("SELECT id, monto, COALESCE(pagado,0) as pagado FROM prestamos WHERE id=?", (row['prestamo_id'],)).fetchone()
    diferencia = nuevo_monto - row['monto']
    nuevo_pagado = prestamo['pagado'] + diferencia
    if nuevo_pagado > prestamo['monto']: return jsonify({'error':'El abono excede la deuda pendiente'}), 400
    db.execute("UPDATE abonos SET monto=? WHERE id=?", (nuevo_monto, id))
    if 'fecha' in data:
        db.execute("UPDATE abonos SET fecha=? WHERE id=?", (data['fecha'], id))
    nuevo_estado = 'pagado' if nuevo_pagado >= prestamo['monto'] else 'pendiente'
    db.execute("UPDATE prestamos SET pagado=?, estado=? WHERE id=?", (nuevo_pagado, nuevo_estado, prestamo['id']))
    db.commit()
    return jsonify({'ok': True, 'saldo': round(prestamo['monto'] - nuevo_pagado, 2), 'estado': nuevo_estado})

@app.route('/api/abonos/<int:id>', methods=['DELETE'])
@login_required
def delete_abono(id):
    db = get_db()
    row = db.execute("SELECT a.id, a.prestamo_id, a.monto FROM abonos a WHERE a.id=?", (id,)).fetchone()
    if not row: return jsonify({'error':'Abono no encontrado'}), 404
    prestamo = db.execute("SELECT id, monto, COALESCE(pagado,0) as pagado FROM prestamos WHERE id=?", (row['prestamo_id'],)).fetchone()
    nuevo_pagado = prestamo['pagado'] - row['monto']
    if nuevo_pagado < 0: nuevo_pagado = 0
    db.execute("DELETE FROM abonos WHERE id=?", (id,))
    nuevo_estado = 'pagado' if nuevo_pagado >= prestamo['monto'] else 'pendiente'
    db.execute("UPDATE prestamos SET pagado=?, estado=? WHERE id=?", (nuevo_pagado, nuevo_estado, prestamo['id']))
    db.commit()
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
#  DATOS COMPLETOS
# ═══════════════════════════════════════════

@app.route('/api/datos/completos', methods=['GET'])
@login_required
def get_datos_completos():
    db = get_db()
    miembros = [dict(r) for r in db.execute("SELECT id, nombre, apodo FROM miembros ORDER BY id").fetchall()]
    fechas_tablas = [r['fecha'] for r in db.execute("SELECT fecha FROM fechas_tablas ORDER BY fecha").fetchall()]

    as_raw = db.execute("SELECT m.nombre, a.fecha, a.valor FROM asistencias a JOIN miembros m ON a.miembro_id = m.id").fetchall()
    asistencias = {}
    for r in as_raw:
        if r['nombre'] not in asistencias: asistencias[r['nombre']] = {}
        asistencias[r['nombre']][r['fecha']] = r['valor']

    bingos = [dict(r) for r in db.execute("SELECT id, fecha, monto, adicional, asistentes FROM bingos ORDER BY fecha").fetchall()]
    for b in bingos:
        dist = db.execute("SELECT miembro_id, recibe, monto_asignado FROM bingo_distribucion WHERE bingo_id=?", (b['id'],)).fetchall()
        b['distribucion'] = [dict(d) for d in dist] if dist else []
    fechas_rifa = [str(i) for i in range(1, 19)]

    ri_raw = db.execute("SELECT m.nombre, r.fecha, r.valor FROM rifas r JOIN miembros m ON r.miembro_id = m.id").fetchall()
    rifas = {}
    for r in ri_raw:
        if r['nombre'] not in rifas: rifas[r['nombre']] = {}
        rifas[r['nombre']][r['fecha']] = r['valor']

    def gah(tipo):
        rows = db.execute("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo=?", (tipo,)).fetchall()
        r = {}
        for row in rows:
            if row['nombre'] not in r: r[row['nombre']] = {}
            r[row['nombre']][row['mes']] = row['valor']
        return r

    cm = [r['clave'] for r in db.execute("SELECT clave FROM cumple_meses ORDER BY orden").fetchall()]
    ca_raw = db.execute("SELECT cm.clave, m.nombre, ca.valor FROM cumple_aportes ca JOIN cumple_meses cm ON ca.cumple_mes_id = cm.id JOIN miembros m ON ca.miembro_id = m.id").fetchall()
    ca = {}
    for r in ca_raw:
        if r['clave'] not in ca: ca[r['clave']] = {}
        ca[r['clave']][r['nombre']] = r['valor']

    cf = [dict(r) for r in db.execute("SELECT m.nombre, cf.fecha FROM cumple_fechas cf JOIN miembros m ON cf.miembro_id = m.id").fetchall()]
    prestamos = [dict(r) for r in db.execute("SELECT p.id, m.nombre, p.monto, p.obs, p.estado, p.fecha, COALESCE(p.pagado,0) as pagado FROM prestamos p JOIN miembros m ON p.miembro_id = m.id ORDER BY p.id").fetchall()]

    return jsonify({
        'miembros': miembros, 'fechasTablas': fechas_tablas, 'asistencias': asistencias,
        'bingos': bingos, 'fechasRifa': fechas_rifa, 'rifas': rifas,
        'ahorroNormal': gah('normal'), 'ahorroCumple': gah('cumple'),
        'ahorroRifa': gah('rifa'), 'cumpleMeses': cm,
        'cumpleAportes': ca, 'cumpleFechas': cf, 'prestamos': prestamos
    })

# ═══════════════════════════════════════════
#  ESTADO DE CUENTA (vista consolidada por miembro)
# ═══════════════════════════════════════════

def _calcular_estado_cuenta(db):
    """Cruza TODAS las tablas del grupo (asistencia, rifa, ahorros, cumpleaños,
    préstamos) para producir un estado de cuenta único por miembro."""
    miembros = [dict(r) for r in db.execute("SELECT id, nombre, apodo FROM miembros ORDER BY id").fetchall()]
    total_fechas_tablas = db.execute("SELECT COUNT(*) c FROM fechas_tablas").fetchone()['c']
    fechas_rifa_list = [str(r['fecha']) for r in db.execute("SELECT fecha FROM fechas_rifa ORDER BY fecha").fetchall()] or [str(i) for i in range(1, 19)]
    total_fechas_rifa = len(fechas_rifa_list)

    asis = {}
    for r in db.execute("SELECT miembro_id, SUM(valor) s, COUNT(*) n FROM asistencias WHERE valor=1 GROUP BY miembro_id"):
        asis[r['miembro_id']] = r['s'] or 0

    # monto_asignado en bingo_distribucion ya refleja el valor real que a cada
    # miembro le toca (auto-calculado o personalizado a mano en el Reparto),
    # así que basta con sumarlo directamente para quienes reciben.
    bingo_dist = {}
    for r in db.execute("SELECT miembro_id, recibe, monto_asignado FROM bingo_distribucion WHERE recibe=TRUE"):
        bingo_dist[r['miembro_id']] = bingo_dist.get(r['miembro_id'], 0) + (r['monto_asignado'] or 0)

    rifa = {}
    rifa_raw = {}
    for r in db.execute("SELECT miembro_id, SUM((valor & 1) + ((valor >> 1) & 1)) s FROM rifas GROUP BY miembro_id"):
        rifa_raw[r['miembro_id']] = r['s'] or 0
    # Cada miembro recibe el total recolectado en su rifa correspondiente (en orden)
    rifa = {}
    for i, m in enumerate(miembros):
        num_rifa = fechas_rifa_list[i] if i < len(fechas_rifa_list) else fechas_rifa_list[0]
        total_recibido = sum(
            ((r['valor'] & 1) + ((r['valor'] >> 1) & 1))
            for r in db.execute("SELECT valor FROM rifas WHERE fecha=?", (num_rifa,)).fetchall()
            if r['valor']
        )
        rifa[m['id']] = total_recibido

    ahorros = {'normal': {}, 'cumple': {}, 'rifa': {}}
    for r in db.execute("SELECT tipo, miembro_id, SUM(valor) s FROM ahorros GROUP BY tipo, miembro_id"):
        ahorros[r['tipo']][r['miembro_id']] = r['s'] or 0

    cumple = {}
    for r in db.execute("SELECT miembro_id, SUM(valor) s FROM cumple_aportes GROUP BY miembro_id"):
        cumple[r['miembro_id']] = r['s'] or 0

    cumple_fecha = {}
    for r in db.execute("SELECT miembro_id, fecha FROM cumple_fechas"):
        cumple_fecha[r['miembro_id']] = r['fecha']

    prest = {}
    for r in db.execute("SELECT miembro_id, SUM(monto - COALESCE(pagado,0)) as s FROM prestamos WHERE estado!='pagado' GROUP BY miembro_id"):
        prest.setdefault(r['miembro_id'], {'pendiente': 0, 'pagado': 0})['pendiente'] = r['s'] or 0
    for r in db.execute("SELECT miembro_id, SUM(COALESCE(pagado,0)) as s FROM prestamos GROUP BY miembro_id"):
        prest.setdefault(r['miembro_id'], {'pendiente': 0, 'pagado': 0})['pagado'] = r['s'] or 0

    resultado = []
    for m in miembros:
        mid = m['id']
        as_ok = asis.get(mid, 0)
        as_pct = round((as_ok / total_fechas_tablas) * 100, 1) if total_fechas_tablas else 0
        aN = ahorros['normal'].get(mid, 0)
        aC = ahorros['cumple'].get(mid, 0)
        aR = rifa.get(mid, 0)  # Calculado de las rifas (cada slot = $1)
        ahorro_total = aN + aC + aR
        p = prest.get(mid, {'pendiente': 0, 'pagado': 0})
        saldo_neto = round(ahorro_total - p['pendiente'], 2)
        estado_general = 'moroso' if p['pendiente'] > 0 and as_pct < 50 else ('con_deuda' if p['pendiente'] > 0 else 'al_dia')
        bingo_ganado = round(bingo_dist.get(mid, 0), 2)
        # Conecta la tabla de Asistencias con la de Bingo: cualquier fecha de
        # "Tablas de Bingo" que el miembro NO tenga marcada como "Sí" (ya sea
        # que la marcó "No" o que nunca la tocó) cuenta como no pagada.
        bingo_debe = max(total_fechas_tablas - as_ok, 0)
        bingo_total = round(bingo_ganado - bingo_debe, 2)
        resultado.append({
            'id': mid, 'nombre': m['nombre'], 'apodo': m['apodo'],
            'asistencia': {'asistidas': as_ok, 'total': total_fechas_tablas, 'pct': as_pct},
            'rifa': {'participaciones': rifa_raw.get(mid, 0), 'totalFechas': total_fechas_rifa},
            'bingo': bingo_total, 'bingoGanado': bingo_ganado, 'bingoDebe': bingo_debe,
            'ahorros': {'normal': round(aN, 2), 'cumple': round(aC, 2), 'rifa': round(aR, 2), 'total': round(ahorro_total, 2)},
            'cumpleAportes': cumple.get(mid, 0),
            'cumpleFecha': cumple_fecha.get(mid),
            'prestamos': {'pendiente': round(p['pendiente'], 2), 'pagado': round(p['pagado'], 2)},
            'saldoNeto': saldo_neto,
            'estadoGeneral': estado_general,
        })
    return resultado

@app.route('/api/estado-cuenta', methods=['GET'])
@login_required
def get_estado_cuenta():
    return jsonify(_calcular_estado_cuenta(get_db()))

@app.route('/api/export/xlsx/estado_cuenta')
@login_required
def export_estado_cuenta_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    db = get_db()
    datos = _calcular_estado_cuenta(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"
    cols = ['N°','Nombre','Asist. (Si/Total)','Asist. %','Rifa (Participaciones)',
            'Bingo Ganado','Tablas No Pagadas ($)','Bingo Neto ($)','Ahorro Normal','Ahorro Cumple','Ahorro Rifa','Total Ahorros',
            'Aportes Cumple','Préstamo Pendiente','Préstamo Pagado','Saldo Neto','Estado']
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    for j, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
    estado_lbl = {'al_dia':'Al día','con_deuda':'Con deuda','moroso':'Moroso'}
    for i, d in enumerate(datos, 1):
        row = [
            i, d['nombre'], f"{d['asistencia']['asistidas']}/{d['asistencia']['total']}", f"{d['asistencia']['pct']}%",
            d['rifa']['participaciones'], d['bingoGanado'], d['bingoDebe'], d['bingo'], d['ahorros']['normal'], d['ahorros']['cumple'], d['ahorros']['rifa'],
            d['ahorros']['total'], d['cumpleAportes'], d['prestamos']['pendiente'], d['prestamos']['pagado'],
            d['saldoNeto'], estado_lbl.get(d['estadoGeneral'], d['estadoGeneral'])
        ]
        for j, val in enumerate(row, 2):
            c = ws.cell(row=i+1, column=j, value=val)
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=i+1, column=1, value=i).alignment = Alignment(horizontal='center')
    for col in ws.columns:
        ml = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 30)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='estado_cuenta_grupo.xlsx')

# ═══════════════════════════════════════════
#  EXPORTAR
# ═══════════════════════════════════════════

def _rows_to_csv(rows, cols):
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(cols)
    for r in rows: w.writerow(r)
    out.seek(0)
    return out

@app.route('/api/export/csv/<tabla>')
@login_required
def export_csv(tabla):
    db = get_db()
    mapping = {
        'miembros': ("SELECT nombre, apodo FROM miembros ORDER BY id", ['nombre','apodo']),
        'asistencias': ("SELECT m.nombre, a.fecha, CASE WHEN a.valor=1 THEN 'Si' ELSE 'No' END FROM asistencias a JOIN miembros m ON a.miembro_id = m.id ORDER BY a.fecha,m.nombre", ['miembro','fecha','asistio']),
        'bingos': ("SELECT fecha, monto FROM bingos ORDER BY fecha", ['fecha','monto']),
        'rifas': ("SELECT m.nombre, r.fecha, r.valor FROM rifas r JOIN miembros m ON r.miembro_id = m.id ORDER BY r.fecha,m.nombre", ['miembro','fecha','cantidad']),
        'ahorro_normal': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='normal' ORDER BY ma.orden,m.nombre", ['miembro','mes','monto']),
        'ahorro_cumple': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='cumple' ORDER BY ma.orden,m.nombre", ['miembro','mes','monto']),
        'ahorro_rifa': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='rifa' ORDER BY ma.orden,m.nombre", ['miembro','mes','monto']),
        'cumple_aportes': ("SELECT cm.clave, m.nombre, ca.valor FROM cumple_aportes ca JOIN cumple_meses cm ON ca.cumple_mes_id = cm.id JOIN miembros m ON ca.miembro_id = m.id ORDER BY cm.orden,m.nombre", ['mes','miembro','cantidad']),
        'cumple_fechas': ("SELECT m.nombre, cf.fecha FROM cumple_fechas cf JOIN miembros m ON cf.miembro_id = m.id ORDER BY cf.fecha", ['miembro','fecha']),
        'prestamos': ("SELECT m.nombre, p.monto, p.obs, p.estado FROM prestamos p JOIN miembros m ON p.miembro_id = m.id ORDER BY p.id", ['miembro','monto','observacion','estado']),
    }
    if tabla not in mapping: return jsonify({'error':'Tabla no encontrada'}), 404
    q, cols = mapping[tabla]
    rows = db.execute(q).fetchall()
    out = _rows_to_csv([tuple(r.values()) for r in rows], cols)
    return Response(out.getvalue(), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename={tabla}.csv'})

@app.route('/api/export/todo')
@login_required
def export_todo():
    import zipfile
    db = get_db()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for key, (q, cols) in {
            'miembros': ("SELECT nombre, apodo FROM miembros ORDER BY id", ['nombre','apodo']),
            'asistencias': ("SELECT m.nombre, a.fecha, CASE WHEN a.valor=1 THEN 'Si' ELSE 'No' END FROM asistencias a JOIN miembros m ON a.miembro_id = m.id ORDER BY a.fecha,m.nombre", ['miembro','fecha','asistio']),
            'bingos': ("SELECT fecha, monto FROM bingos ORDER BY fecha", ['fecha','monto']),
        'rifas': ("SELECT m.nombre, r.fecha as mes, r.valor FROM rifas r JOIN miembros m ON r.miembro_id = m.id ORDER BY r.fecha,m.nombre", ['miembro','mes','cantidad']),
            'ahorro_normal': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='normal' ORDER BY ma.orden,m.nombre", ['miembro','mes','monto']),
            'ahorro_cumple': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='cumple' ORDER BY ma.orden,m.nombre", ['miembro','mes','monto']),
            'ahorro_rifa': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='rifa' ORDER BY ma.orden,m.nombre", ['miembro','mes','monto']),
            'cumple_aportes': ("SELECT cm.clave, m.nombre, ca.valor FROM cumple_aportes ca JOIN cumple_meses cm ON ca.cumple_mes_id = cm.id JOIN miembros m ON ca.miembro_id = m.id ORDER BY cm.orden,m.nombre", ['mes','miembro','cantidad']),
            'cumple_fechas': ("SELECT m.nombre, cf.fecha FROM cumple_fechas cf JOIN miembros m ON cf.miembro_id = m.id ORDER BY cf.fecha", ['miembro','fecha']),
            'prestamos': ("SELECT m.nombre, p.monto, p.obs, p.estado FROM prestamos p JOIN miembros m ON p.miembro_id = m.id ORDER BY p.id", ['miembro','monto','observacion','estado']),
        }.items():
            rows = db.execute(q).fetchall()
            out = _rows_to_csv([tuple(r.values()) for r in rows], cols)
            zf.writestr(f'{key}.csv', out.getvalue())
    buf.seek(0)
    return send_file(buf, mimetype='application/zip', as_attachment=True, download_name='grupo_portoviejo_completo.zip')

# ═══════════════════════════════════════════
#  EXPORTAR EXCEL
# ═══════════════════════════════════════════

def _rows_to_xlsx(rows, cols):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    for j, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = Alignment(horizontal='center', vertical='center')
    for col in ws.columns:
        ml = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 45)
    return wb

_TABLA_MAPPING = {
    'miembros': ("SELECT nombre, apodo FROM miembros ORDER BY id", ['Nombre','Apodo']),
    'asistencias': ("SELECT m.nombre, a.fecha, CASE WHEN a.valor=1 THEN 'Si' ELSE 'No' END FROM asistencias a JOIN miembros m ON a.miembro_id = m.id ORDER BY a.fecha,m.nombre", ['Miembro','Fecha','Asistió']),
    'bingos': ("SELECT fecha, monto FROM bingos ORDER BY fecha", ['Fecha','Monto']),
    'rifas': ("SELECT m.nombre, r.fecha as mes, r.valor FROM rifas r JOIN miembros m ON r.miembro_id = m.id ORDER BY r.fecha,m.nombre", ['Miembro','Mes','Cantidad']),
    'ahorro_normal': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='normal' ORDER BY ma.orden,m.nombre", ['Miembro','Mes','Monto']),
    'ahorro_cumple': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='cumple' ORDER BY ma.orden,m.nombre", ['Miembro','Mes','Monto']),
    'ahorro_rifa': ("SELECT m.nombre, ma.nombre as mes, a.valor FROM ahorros a JOIN miembros m ON a.miembro_id = m.id JOIN meses_ahorro ma ON a.mes_id = ma.id WHERE a.tipo='rifa' ORDER BY ma.orden,m.nombre", ['Miembro','Mes','Monto']),
    'cumple_aportes': ("SELECT cm.clave, m.nombre, ca.valor FROM cumple_aportes ca JOIN cumple_meses cm ON ca.cumple_mes_id = cm.id JOIN miembros m ON ca.miembro_id = m.id ORDER BY cm.orden,m.nombre", ['Mes','Miembro','Cantidad']),
    'cumple_fechas': ("SELECT m.nombre, cf.fecha FROM cumple_fechas cf JOIN miembros m ON cf.miembro_id = m.id ORDER BY cf.fecha", ['Miembro','Fecha']),
    'prestamos': ("SELECT m.nombre, p.monto, p.obs, p.estado FROM prestamos p JOIN miembros m ON p.miembro_id = m.id ORDER BY p.id", ['Miembro','Monto','Observación','Estado']),
}

@app.route('/api/export/xlsx/<tabla>')
@login_required
def export_xlsx(tabla):
    if tabla not in _TABLA_MAPPING:
        return jsonify({'error':'Tabla no encontrada'}), 404
    q, cols = _TABLA_MAPPING[tabla]
    rows = get_db().execute(q).fetchall()
    wb = _rows_to_xlsx([tuple(r.values()) for r in rows], cols)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{tabla}.xlsx')

@app.route('/api/export/xlsx/todo')
@login_required
def export_xlsx_todo():
    import zipfile
    db = get_db()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for key, (q, cols) in _TABLA_MAPPING.items():
            rows = db.execute(q).fetchall()
            wb = _rows_to_xlsx([tuple(r.values()) for r in rows], cols)
            xl_buf = io.BytesIO()
            wb.save(xl_buf)
            xl_buf.seek(0)
            zf.writestr(f'{key}.xlsx', xl_buf.getvalue())
    buf.seek(0)
    return send_file(buf, mimetype='application/zip', as_attachment=True, download_name='grupo_portoviejo_completo_xlsx.zip')

# ═══════════════════════════════════════════
#  REPORTE DE BINGO (profesional)
# ═══════════════════════════════════════════

@app.route('/api/export/xlsx/bingo_reporte')
@login_required
def export_bingo_reporte():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    db = get_db()
    bingos = [dict(r) for r in db.execute("SELECT id, fecha, monto, adicional, asistentes FROM bingos ORDER BY fecha").fetchall()]
    miembros = [dict(r) for r in db.execute("SELECT id, nombre FROM miembros ORDER BY id").fetchall()]
    asistencias_raw = [dict(r) for r in db.execute("SELECT miembro_id, fecha FROM asistencias WHERE valor=1").fetchall()]

    # Mapa de asistencia: fecha -> set de miembro_id
    asist_map = {}
    for a in asistencias_raw:
        asist_map.setdefault(a['fecha'], set()).add(a['miembro_id'])

    wb = Workbook()

    # ── Estilos ──
    title_font = Font(name='Calibri', size=16, bold=True, color='0F172A')
    subtitle_font = Font(name='Calibri', size=10, color='64748B')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    accent_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    green_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    light_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    total_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    data_font = Font(name='Calibri', size=11, color='1E293B')
    bold_font = Font(name='Calibri', size=11, bold=True, color='1E293B')
    money_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    # ═══════════════════════════════════════
    # HOJA 1: Resumen
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen de Bingos"

    # Título
    ws1.merge_cells('A1:E1')
    c = ws1.cell(row=1, column=1, value='Reporte de Bingos — Grupo Calle Portoviejo')
    c.font = title_font; c.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells('A2:E2')
    c = ws1.cell(row=2, column=1, value=f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M")}')
    c.font = subtitle_font; c.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[2].height = 18

    # Encabezados
    headers1 = ['#', 'Fecha', 'Monto Total', 'Asistentes', 'Monto por Persona']
    for j, h in enumerate(headers1, 1):
        c = ws1.cell(row=4, column=j, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = thin_border
    ws1.row_dimensions[4].height = 24

    # Datos
    gran_total = 0
    for i, b in enumerate(bingos, 1):
        row_num = i + 4
        asistentes = b.get('asistentes') or len(miembros)
        denom = asistentes if asistentes > 0 else 1
        por_persona = round(b['monto'] / denom, 2)
        gran_total += b['monto']

        vals = [i, b['fecha'], b['monto'], asistentes, por_persona]
        for j, v in enumerate(vals, 1):
            c = ws1.cell(row=row_num, column=j, value=v)
            c.font = data_font; c.alignment = center; c.border = thin_border
            if j == 1: c.alignment = center
            if j == 3 or j == 5:
                c.number_format = money_fmt
        if i % 2 == 0:
            for j in range(1, 6):
                ws1.cell(row=row_num, column=j).fill = light_fill

    # Fila de total
    total_row = len(bingos) + 5
    ws1.merge_cells(f'A{total_row}:B{total_row}')
    c = ws1.cell(row=total_row, column=1, value='TOTAL GENERAL')
    c.font = bold_font; c.alignment = center; c.fill = total_fill
    for j in range(1, 6):
        ws1.cell(row=total_row, column=j).border = thin_border
        ws1.cell(row=total_row, column=j).fill = total_fill
    c = ws1.cell(row=total_row, column=3, value=gran_total)
    c.font = bold_font; c.number_format = money_fmt; c.alignment = center

    # Anchos
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 22

    # ═══════════════════════════════════════
    # HOJA 2: Detalle por Miembro
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("Detalle por Miembro")

    ws2.merge_cells(f'A1:{get_column_letter(len(bingos)+2)}1')
    c = ws2.cell(row=1, column=1, value='Distribución de Bingos por Miembro')
    c.font = title_font; c.alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells(f'A2:{get_column_letter(len(bingos)+2)}2')
    c = ws2.cell(row=2, column=1, value='Monto distribuido equitativamente entre todos los miembros registrados')
    c.font = subtitle_font

    # Encabezados: #, Nombre, [fechas...], Total
    h2 = ['#', 'Miembro'] + [b['fecha'] for b in bingos] + ['Total Recibido']
    for j, h in enumerate(h2, 1):
        c = ws2.cell(row=4, column=j, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = thin_border

    # Matriz de montos por miembro por bingo
    montos_por_miembro = {}
    for m in miembros:
        montos_por_miembro[m['nombre']] = {}
        total_miembro = 0
        for b in bingos:
            asistentes = b.get('asistentes') or len(miembros)
            denom = asistentes if asistentes > 0 else 1
            por_persona = round(b['monto'] / denom, 2)
            montos_por_miembro[m['nombre']][b['fecha']] = por_persona
            total_miembro += por_persona
        montos_por_miembro[m['nombre']]['_total'] = round(total_miembro, 2)

    for i, m in enumerate(miembros, 1):
        row_num = i + 4
        ws2.cell(row=row_num, column=1, value=i).font = data_font
        ws2.cell(row=row_num, column=1).alignment = center
        ws2.cell(row=row_num, column=1).border = thin_border

        c = ws2.cell(row=row_num, column=2, value=m['nombre'])
        c.font = bold_font; c.alignment = left; c.border = thin_border

        for j, b in enumerate(bingos, 3):
            val = montos_por_miembro[m['nombre']][b['fecha']]
            c = ws2.cell(row=row_num, column=j, value=val if val > 0 else 0)
            c.font = data_font; c.alignment = center; c.border = thin_border
            c.number_format = money_fmt
            if val > 0:
                c.fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')

        # Total por miembro
        col_total = len(bingos) + 3
        c = ws2.cell(row=row_num, column=col_total, value=montos_por_miembro[m['nombre']]['_total'])
        c.font = bold_font; c.alignment = center; c.border = thin_border
        c.number_format = money_fmt
        c.fill = green_fill
        c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

        if i % 2 == 0:
            for j in range(1, col_total + 1):
                cell = ws2.cell(row=row_num, column=j)
                if not cell.fill or cell.fill.start_color.index == '00000000':
                    cell.fill = light_fill

    # Fila total
    total_row2 = len(miembros) + 5
    ws2.cell(row=total_row2, column=1, value='').border = thin_border
    c = ws2.cell(row=total_row2, column=2, value='TOTALES')
    c.font = bold_font; c.alignment = center; c.fill = total_fill
    c.border = thin_border
    gran_por_miembro = 0
    for j, b in enumerate(bingos, 3):
        suma = sum(montos_por_miembro[m['nombre']][b['fecha']] for m in miembros)
        c = ws2.cell(row=total_row2, column=j, value=round(suma, 2))
        c.font = bold_font; c.alignment = center; c.border = thin_border
        c.number_format = money_fmt; c.fill = total_fill
        gran_por_miembro += suma
    col_t = len(bingos) + 3
    c = ws2.cell(row=total_row2, column=col_t, value=round(gran_por_miembro, 2))
    c.font = bold_font; c.alignment = center; c.border = thin_border
    c.number_format = money_fmt; c.fill = total_fill

    # Anchos columna hoja 2
    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 28
    for j in range(3, col_t + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 14

    # ═══════════════════════════════════════
    # HOJA 3: Resumen por Miembro
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("Totales por Miembro")

    ws3.merge_cells('A1:C1')
    c = ws3.cell(row=1, column=1, value='Totales Recibidos por Miembro')
    c.font = title_font
    ws3.row_dimensions[1].height = 30

    h3 = ['#', 'Miembro', 'Total Recibido en Bingos']
    for j, h in enumerate(h3, 1):
        c = ws3.cell(row=3, column=j, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = thin_border

    miembros_ordenados = sorted(miembros, key=lambda m: montos_por_miembro[m['nombre']]['_total'], reverse=True)
    for i, m in enumerate(miembros_ordenados, 1):
        row_num = i + 3
        ws3.cell(row=row_num, column=1, value=i).font = data_font
        ws3.cell(row=row_num, column=1).alignment = center
        ws3.cell(row=row_num, column=1).border = thin_border

        c = ws3.cell(row=row_num, column=2, value=m['nombre'])
        c.font = bold_font; c.alignment = left; c.border = thin_border

        total = montos_por_miembro[m['nombre']]['_total']
        c = ws3.cell(row=row_num, column=3, value=total)
        c.font = Font(name='Calibri', size=11, bold=True, color='059669')
        c.alignment = center; c.border = thin_border
        c.number_format = money_fmt

        if i % 2 == 0:
            ws3.cell(row=row_num, column=1).fill = light_fill
            ws3.cell(row=row_num, column=2).fill = light_fill
            ws3.cell(row=row_num, column=3).fill = light_fill

    total_row3 = len(miembros_ordenados) + 4
    ws3.cell(row=total_row3, column=1, value='').border = thin_border
    c = ws3.cell(row=total_row3, column=2, value='TOTAL GENERAL')
    c.font = bold_font; c.alignment = center; c.fill = total_fill; c.border = thin_border
    c = ws3.cell(row=total_row3, column=3, value=round(gran_por_miembro, 2))
    c.font = bold_font; c.alignment = center; c.fill = total_fill; c.border = thin_border
    c.number_format = money_fmt

    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 28
    ws3.column_dimensions['C'].width = 26

    # ── Guardar ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='reporte_bingos.xlsx')

# ═══════════════════════════════════════════
#  IMPORTAR DESDE EXCEL
# ═══════════════════════════════════════════

@app.route('/api/import/excel', methods=['POST'])
@login_required
def import_excel():
    if 'file' not in request.files:
        return jsonify({'error':'No se envió archivo'}), 400
    f = request.files['file']
    if not f.filename.endswith('.xlsx'):
        return jsonify({'error':'Solo archivos .xlsx'}), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({'error':'openpyxl no instalado'}), 500

    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify({'error':f'Error al leer Excel: {str(e)}'}), 400

    db = get_db()
    resumen = {'miembros':0, 'asistencias':0, 'bingos':0, 'rifas':0, 'ahorros':0, 'cumple':0, 'prestamos':0}
    errores = []
    hojas_detectadas = []

    for ws in wb.worksheets:
        name = ws.title.strip().upper()
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw: continue

        # ── Limpiar filas vacías al inicio y final ──
        while rows_raw and all(c is None or (isinstance(c, str) and c.strip() == '') for c in rows_raw[0]):
            rows_raw.pop(0)
        while rows_raw and all(c is None or (isinstance(c, str) and c.strip() == '') for c in rows_raw[-1]):
            rows_raw.pop()
        if not rows_raw: continue

        # ── Detectar y saltar filas de título (texto largo en col A, sin datos en cols siguientes) ──
        def _es_titulo(row):
            first = str(row[0]).strip() if row[0] else ''
            if not first: return False
            # Si la primera celda es muy larga (>25) y el resto están vacías → título
            if len(first) > 25:
                return True
            # Si coincide con patrones comunes de título
            if first.upper() in ('GRUPO CALLE PORTOVIEJO', 'GRUPO CALLE PORTOVIEJO 2026', 'GRUPO', 'CALLE', 'PORTOVIEJO'):
                return True
            return False

        while rows_raw and _es_titulo(rows_raw[0]):
            rows_raw.pop(0)
        while rows_raw and all(c is None or (isinstance(c, str) and c.strip() == '') for c in rows_raw[0]):
            rows_raw.pop(0)
        if not rows_raw: continue

        rows_list = rows_raw
        hojas_detectadas.append(name)

        # ── MIEMBROS ──
        if 'MIEMBRO' in name:
            start = 1 if str(rows_list[0][0]).strip().upper() in ('NOMBRE', 'NOMBRE COMPLETO', 'NOMBRE COMPLETO DEL MIEMBRO', 'MIEMBRO', 'APELLIDOS Y NOMBRES') else 0
            for row in rows_list[start:]:
                nom = str(row[0]).strip().upper() if row[0] else ''
                apo = str(row[1]).strip() if len(row)>1 and row[1] else ''
                if nom and nom not in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', 'APODO', ''):
                    try:
                        db.execute("INSERT INTO miembros (nombre, apodo) VALUES (?, ?)", (nom, apo))
                        resumen['miembros'] += 1
                    except psycopg2.IntegrityError: pass
            db.commit()
            continue

        # ── ASISTENCIAS / TABLAS ──
        if 'ASISTENCIA' in name or name in ('TABLAS', 'ASIST', 'ASIS'):
            # Buscar fila de encabezados (donde la 1ra col NO Parece nombre de miembro)
            hdr_idx = 0
            for i, row in enumerate(rows_list):
                first = str(row[0]).strip().upper() if row[0] else ''
                if first in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', 'APELLIDOS Y NOMBRES', ''):
                    hdr_idx = i
                    break
            headers = [str(c) for c in rows_list[hdr_idx]] if rows_list[hdr_idx] else []
            for row in rows_list[hdr_idx+1:]:
                nom = str(row[0]).strip().upper() if row[0] else ''
                if not nom or nom in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', ''): continue
                miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (nom,)).fetchone()
                if not miembro: continue
                for j, fecha_raw in enumerate(headers[1:], 1):
                    if j >= len(row): break
                    fecha = str(fecha_raw).strip() if fecha_raw else ''
                    if not fecha or fecha.upper() in ('NONE', 'NINGUNO', ''): continue
                    val = 1 if row[j] and str(row[j]).strip() in ('1','Si','SI','si','x','X','✓','✔','SÍ') else 0
                    try:
                        db.execute("INSERT INTO fechas_tablas (fecha) VALUES (?) ON CONFLICT (fecha) DO NOTHING", (fecha,))
                        db.execute("INSERT INTO asistencias (miembro_id, fecha, valor) VALUES (?, ?, ?) ON CONFLICT(miembro_id, fecha) DO UPDATE SET valor=excluded.valor",
                                   (miembro['id'], fecha, val))
                        resumen['asistencias'] += 1
                    except: pass
            db.commit()
            continue

        # ── BINGO ──
        if 'BINGO' in name:
            start = 1 if rows_list[0][0] and str(rows_list[0][0]).strip().upper() in ('FECHA', 'FECHAS') else 0
            for row in rows_list[start:]:
                if not row[0] or not row[1]: continue
                fecha = str(row[0]).strip()
                if fecha.upper() in ('FECHA', 'FECHAS', 'NONE', ''): continue
                try:
                    db.execute("INSERT INTO bingos (fecha, monto) VALUES (?, ?) ON CONFLICT (fecha) DO NOTHING",
                               (fecha, float(row[1])))
                    resumen['bingos'] += 1
                except: pass
            db.commit()
            continue

        # ── RIFA ──
        if 'RIFA' in name:
            hdr_idx = 0
            for i, row in enumerate(rows_list):
                first = str(row[0]).strip().upper() if row[0] else ''
                if first in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', ''):
                    hdr_idx = i
                    break
            headers = [str(c) for c in rows_list[hdr_idx]] if rows_list[hdr_idx] else []
            for row in rows_list[hdr_idx+1:]:
                nom = str(row[0]).strip().upper() if row[0] else ''
                if not nom or nom in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', ''): continue
                miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (nom,)).fetchone()
                if not miembro: continue
                for j, fecha_raw in enumerate(headers[1:], 1):
                    if j >= len(row): break
                    fecha = str(fecha_raw).strip() if fecha_raw else ''
                    if not fecha or fecha.upper() in ('NONE', 'TOTAL', ''): continue
                    try: val = int(float(str(row[j]))) if row[j] else 0
                    except: val = 0
                    try:
                        db.execute("INSERT INTO fechas_rifa (fecha) VALUES (?) ON CONFLICT (fecha) DO NOTHING", (fecha,))
                        db.execute("INSERT INTO rifas (miembro_id, fecha, valor) VALUES (?, ?, ?) ON CONFLICT(miembro_id, fecha) DO UPDATE SET valor=excluded.valor",
                                   (miembro['id'], fecha, val))
                        resumen['rifas'] += 1
                    except: pass
            db.commit()
            continue

        # ── AHORROS ──
        if 'AHORRO' in name:
            hdr_idx = 0
            for i, row in enumerate(rows_list):
                first = str(row[0]).strip().upper() if row[0] else ''
                if first in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', ''):
                    hdr_idx = i
                    break
            headers = [str(c).strip().upper() for c in rows_list[hdr_idx]] if rows_list[hdr_idx] else []
            tipo = 'normal'
            if 'CUMPLE' in name: tipo = 'cumple'
            if 'RIFA' in name: tipo = 'rifa'
            for row in rows_list[hdr_idx+1:]:
                nom = str(row[0]).strip().upper() if row[0] else ''
                if not nom or nom in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', ''): continue
                miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (nom,)).fetchone()
                if not miembro: continue
                for j, h in enumerate(headers[1:], 1):
                    if j >= len(row): break
                    if h in ('TOTAL', 'SUMA', ''): continue
                    mes = h.capitalize()
                    if mes not in ('Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'): continue
                    try: val = float(row[j]) if row[j] else 0
                    except: val = 0
                    if val:
                        mes_row = db.execute("SELECT id FROM meses_ahorro WHERE nombre=?", (mes,)).fetchone()
                        if mes_row:
                            db.execute("INSERT INTO ahorros (tipo, miembro_id, mes_id, valor) VALUES (?, ?, ?, ?) ON CONFLICT(tipo, miembro_id, mes_id) DO UPDATE SET valor=excluded.valor",
                                       (tipo, miembro['id'], mes_row['id'], val))
                            resumen['ahorros'] += 1
            db.commit()
            continue

        # ── CUMPLEAÑOS (APORTES) ──
        if 'CUMPLE' in name:
            hdr_idx = 0
            for i, row in enumerate(rows_list):
                first = str(row[0]).strip().upper() if row[0] else ''
                if first in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', ''):
                    hdr_idx = i
                    break
            headers = [str(c).strip() for c in rows_list[hdr_idx]] if rows_list[hdr_idx] else []
            for row in rows_list[hdr_idx+1:]:
                nom = str(row[0]).strip().upper() if row[0] else ''
                if not nom or nom in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', ''): continue
                miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (nom,)).fetchone()
                if not miembro: continue
                for j, h in enumerate(headers[1:], 1):
                    if j >= len(row): break
                    if h.upper() in ('TOTAL', 'SUMA', ''): continue
                    try: val = int(float(str(row[j]))) if row[j] else 0
                    except: val = 0
                    if val:
                        cm = db.execute("SELECT id FROM cumple_meses WHERE clave=?", (h.strip(),)).fetchone()
                        if cm:
                            db.execute("INSERT INTO cumple_aportes (cumple_mes_id, miembro_id, valor) VALUES (?, ?, ?) ON CONFLICT(cumple_mes_id, miembro_id) DO UPDATE SET valor=excluded.valor",
                                       (cm['id'], miembro['id'], val))
                            resumen['cumple'] += 1
            db.commit()
            continue

        # ── PRESTAMOS ──
        if 'PRESTAMO' in name:
            start = 1 if rows_list[0][0] and str(rows_list[0][0]).strip().upper() in ('NOMBRE', 'MIEMBRO') else 0
            for row in rows_list[start:]:
                nom = str(row[0]).strip().upper() if row[0] else ''
                if not nom or nom in ('NOMBRE', 'NOMBRE COMPLETO', 'MIEMBRO', 'MONTO', 'OBSERVACIÓN', ''): continue
                try: monto = float(row[1]) if len(row)>1 and row[1] else 0
                except: monto = 0
                if not monto: continue
                obs = str(row[2]).strip() if len(row)>2 and row[2] else ''
                miembro = db.execute("SELECT id FROM miembros WHERE nombre=?", (nom,)).fetchone()
                if miembro:
                    db.execute("INSERT INTO prestamos (miembro_id, monto, obs) VALUES (?, ?, ?)", (miembro['id'], monto, obs))
                    resumen['prestamos'] += 1
            db.commit()
            continue

    db.commit()
    return jsonify({'ok': True, 'resumen': resumen, 'errores': errores, 'hojas': hojas_detectadas})

# ═══════════════════════════════════════════
#  SPA
# ═══════════════════════════════════════════

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

# ═══════════════════════════════════════════
#  INIT
# ═══════════════════════════════════════════

# Se ejecuta siempre al importar el módulo (tanto con "python app.py" como
# con Gunicorn en Render), para asegurar que la BD y la config existan.
init_db()
init_config()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print("=" * 50)
    print("  Grupo Calle Portoviejo - Servidor SQL")
    print("  DB: PostgreSQL (DATABASE_URL)")
    print(f"  http://localhost:{port}")
    print("  Password default: admin123")
    print("=" * 50)
    # debug=False siempre en producción (Render). Para desarrollo local
    # puedes cambiarlo temporalmente a True.
    app.run(host='0.0.0.0', port=port, debug=False)
